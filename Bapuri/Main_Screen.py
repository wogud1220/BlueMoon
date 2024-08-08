import sys, re
from datetime import datetime
import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox, QLineEdit, QToolButton, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QSpacerItem, QSizePolicy, QFileDialog, QSplitter
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QFont
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 시장조사 때문에 느려진듯
# 엑셀 파일 경로 설정
#data_path_purchase = './2024_식자재매입_test.xlsx' # 매출매입 페이지: 매입표
data_path_sales = './2024_식자재매출_test.xlsx' # 매출매입 페이지: 매출표
data_path_cost_ratio = './2024_식자재매출_test.xlsx' # 원가율 페이지: 원가윮표

# 매입, 매출 구성 함수
class ExcelViewer(QWidget):
    def __init__(self, filePath, label):
        super().__init__()
        self.filePath = filePath
        self.label = label
        self.initUI()
        self.loadExcelData(self.filePath)

    # 매출, 매입 페이지의 UI 초기화 함수
    def initUI(self):
        self.layout = QVBoxLayout()

        self.labelWidget = QLabel(self.label, self)
        self.labelWidget.setAlignment(Qt.AlignCenter)

        self.tableWidget = QTableWidget(self)
        self.layout.addWidget(self.labelWidget)
        self.layout.addWidget(self.tableWidget)

        buttonLayout = QHBoxLayout()

        self.btnAddRow = QPushButton('Add Row', self)
        self.btnAddRow.clicked.connect(self.addRow)
        buttonLayout.addWidget(self.btnAddRow)

        self.btnRemoveRow = QPushButton('Remove Selected Row', self)
        self.btnRemoveRow.clicked.connect(self.removeRow)
        buttonLayout.addWidget(self.btnRemoveRow)

        self.btnSave = QPushButton('Save Changes', self)
        self.btnSave.clicked.connect(self.saveChanges)
        buttonLayout.addWidget(self.btnSave)

        self.layout.addLayout(buttonLayout)
        self.setLayout(self.layout)

    # 엑셀 데이터 로드 함수
    def loadExcelData(self, filePath):
        self.df = pd.read_excel(filePath)
        if self.df is not None:
            self.tableWidget.setRowCount(self.df.shape[0])
            self.tableWidget.setColumnCount(self.df.shape[1])
            self.tableWidget.setHorizontalHeaderLabels(self.df.columns)

            for i in range(self.df.shape[0]):
                for j in range(self.df.shape[1]):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(str(self.df.iat[i, j])))

    # 변경 사항 저장 함수
    def saveChanges(self):
        new_data = []
        for i in range(self.tableWidget.rowCount()):
            row_data = []
            for j in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(i, j)
                if item:
                    row_data.append(item.text())
                else:
                    row_data.append(None)
            new_data.append(row_data)

        new_df = pd.DataFrame(new_data, columns=self.df.columns)

        for column in self.df.columns:
            if pd.api.types.is_numeric_dtype(self.df[column]):
                new_df[column] = pd.to_numeric(new_df[column], errors='coerce')
            else:
                new_df[column] = new_df[column].astype(self.df[column].dtype)

        self.df = new_df
        self.df.to_excel(self.filePath, index=False)
        QMessageBox.information(self, 'Success', 'Changes saved successfully!')

    # 행 추가 함수
    def addRow(self):
        rowPosition = self.tableWidget.rowCount()
        self.tableWidget.insertRow(rowPosition)

    # 선택된 행 삭제 함수
    def removeRow(self):
        currentRow = self.tableWidget.currentRow()
        if currentRow >= 0:
            self.tableWidget.removeRow(currentRow)
        else:
            QMessageBox.warning(self, 'Warning', 'Please select a row to remove.')

# 읽기 전용 엑셀 뷰어 클래스 정의
class ReadOnlyExcelViewer(ExcelViewer):
    def initUI(self):
        self.layout = QVBoxLayout()

        self.labelWidget = QLabel(self.label, self)
        self.labelWidget.setAlignment(Qt.AlignCenter)

        self.tableWidget = QTableWidget(self)
        self.layout.addWidget(self.labelWidget)
        self.layout.addWidget(self.tableWidget)

        self.setLayout(self.layout)

    def saveChanges(self):
        pass

    def addRow(self):
        pass

    def removeRow(self):
        pass

# 매입, 매출 호출
class DualExcelViewer(QWidget):
    def __init__(self, leftFilePath, rightFilePath):
        super().__init__()
        self.initUI(leftFilePath, rightFilePath)

    def initUI(self, leftFilePath, rightFilePath):
        self.layout = QVBoxLayout()

        self.splitter = QSplitter(Qt.Horizontal)
        self.splitter.setStyleSheet("QSplitter::handle { background-color: gray; }")

        self.leftViewer = ExcelViewer(leftFilePath, '매입')
        self.rightViewer = ExcelViewer(rightFilePath, '매출')

        self.splitter.addWidget(self.leftViewer)
        self.splitter.addWidget(self.rightViewer)

        self.splitter.splitterMoved.connect(self.adjustFontSize)

        self.layout.addWidget(self.splitter)
        self.setLayout(self.layout)

        self.adjustFontSize()

    def adjustFontSize(self):
        leftWidth = self.splitter.widget(0).width()
        rightWidth = self.splitter.widget(1).width()

        leftFontSize = max(10, leftWidth // 10)
        rightFontSize = max(10, rightWidth // 10)

        self.leftViewer.labelWidget.setFont(self.setFontSize(leftFontSize))
        self.rightViewer.labelWidget.setFont(self.setFontSize(rightFontSize))

    def setFontSize(self, size):
        font = self.leftViewer.labelWidget.font()
        font.setPointSize(size)
        return font


class Ui_StackedWidget(object):
    def setupUi(self, StackedWidget):
        StackedWidget.setObjectName("StackedWidget")
        StackedWidget.resize(1200, 900)
        # 페이지 구성 설정
        self.create_pages(StackedWidget)
        self.retranslateUi(StackedWidget)
        QtCore.QMetaObject.connectSlotsByName(StackedWidget)
    # 각 페이지 생성 선언
    def create_pages(self, StackedWidget):
        
        self.page = QtWidgets.QWidget()
        self.page.setObjectName("page")
        self.setup_home_page(self.page, StackedWidget)
        StackedWidget.addWidget(self.page)

        # 나머지 페이지들도 각 함수 내에서 self.page_x, 객체 가지고 있음
        self.page_2 = QtWidgets.QWidget()
        self.page_2.setObjectName("page_2")
        self.setup_purchase_sale_page(self.page_2, StackedWidget)
        StackedWidget.addWidget(self.page_2)

        self.setup_market_research_page(StackedWidget)  # 여기에서 self.page_3을 설정하므로 인수를 하나로 줄입니다.

        self.setup_cost_rate_page(StackedWidget)  # 여기에서 self.page_4을 설정하므로 인수를 하나로 줄입니다.

        self.setup_order_form_page(StackedWidget)  # 여기에서 self.page_5을 설정하므로 인수를 하나로 줄입니다.

        # 데이터 로드, 시장 조사 테이블 로드 안됨
        # self.df = None
        # self.excel_data = None
        # self.load_excel()
        # self.load_order_table(self)


    # 메인 페이지 생성
    def setup_home_page(self, page, StackedWidget):
        self.homeButton1 = QtWidgets.QPushButton(page)
        self.homeButton1.setText("홈(사진)")
        self.homeButton1.setAutoDefault(False)
        self.homeButton1.clicked.connect(lambda: (StackedWidget.setCurrentIndex(0),))
        # self.save_current_data_to_excel()))
        font = QtGui.QFont()
        font.setPointSize(25)

        self.toolButton = QToolButton(page)
        self.toolButton.setFont(font)
        self.toolButton.setStyleSheet("color:white; background-color:rgb(230,100,30);")
        self.toolButton.setObjectName("toolButton")
        self.toolButton.setText("매입매출")
        self.toolButton.clicked.connect(lambda: (StackedWidget.setCurrentIndex(1),)) 
        #self.save_current_data_to_excel()))

        self.toolButton_2 = QToolButton(page)
        self.toolButton_2.setFont(font)
        self.toolButton_2.setObjectName("toolButton_2")
        self.toolButton_2.setText("시장조사")
        self.toolButton_2.clicked.connect(lambda: (StackedWidget.setCurrentIndex(2), ))
        #self.save_current_data_to_excel()))

        self.toolButton_3 = QToolButton(page)
        self.toolButton_3.setFont(font)
        self.toolButton_3.setObjectName("toolButton_3")
        self.toolButton_3.setText("원가율")
        self.toolButton_3.clicked.connect(lambda: (StackedWidget.setCurrentIndex(3),)) 
        #self.save_current_data_to_excel()))

        self.toolButton_4 = QToolButton(page)
        self.toolButton_4.setFont(font)
        self.toolButton_4.setObjectName("toolButton_4")
        self.toolButton_4.setText("발주서")
        self.toolButton_4.clicked.connect(lambda: (StackedWidget.setCurrentIndex(4),))
        #self.save_current_data_to_excel()))

        font.setPointSize(40)
        self.label_2 = QtWidgets.QLabel(page)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_2.setText("      바푸리애프앤비")

        self.pageLayout = QVBoxLayout(page)
        self.pageLayout.addWidget(self.homeButton1, alignment=QtCore.Qt.AlignLeft)
        self.pageLayout.addWidget(self.label_2, alignment=QtCore.Qt.AlignCenter)

        self.buttonLayout = QHBoxLayout()
        self.buttonLayout.addWidget(self.toolButton)
        self.buttonLayout.addWidget(self.toolButton_2)
        self.buttonLayout.addWidget(self.toolButton_3)
        self.buttonLayout.addWidget(self.toolButton_4)
        self.pageLayout.addLayout(self.buttonLayout)
        StackedWidget.addWidget(page)

    # 페이지 2 - 매입매출 페이지 설정
    def setup_purchase_sale_page(self, page, StackedWidget):
        self.page_2 = QtWidgets.QWidget()
        self.page_2.setObjectName("page_2")

        self.homeButton2 = QtWidgets.QPushButton(self.page_2)
        self.homeButton2.setText("홈(사진)")
        self.homeButton2.setAutoDefault(False)
        self.homeButton2.clicked.connect(lambda: StackedWidget.setCurrentIndex(0))

        self.page2Layout = QVBoxLayout(self.page_2)
        self.page2Layout.addWidget(self.homeButton2, alignment=QtCore.Qt.AlignLeft)

        # DualExcelViewer 추가
        self.dualExcelViewer = DualExcelViewer(#data_path_purchase
            './매입.xlsx', data_path_sales)
        self.page2Layout.addWidget(self.dualExcelViewer)
        StackedWidget.addWidget(self.page_2)

    # 페이지 3 - 시장조사 페이지 설정
    
    def setup_market_research_page(self, StackedWidget):

        self.page_3 = QtWidgets.QWidget()
        self.page_3.setObjectName("page_3")

        self.homeButton3 = QtWidgets.QPushButton(self.page_3)
        self.homeButton3.setText("홈(사진)")
        self.homeButton3.setAutoDefault(False)
        self.homeButton3.clicked.connect(lambda: (StackedWidget.setCurrentIndex(0)))
                                         #, self.save_current_data_to_excel()))

        font = QtGui.QFont()
        font.setPointSize(25)
        self.label_4 = QtWidgets.QLabel(self.page_3)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_4.setText("      시장조사 페이지")

        self.lineEdit = QtWidgets.QLineEdit(self.page_3)
        self.lineEdit.setPlaceholderText("음식 이름")
        font_search_food = self.lineEdit.font()
        font_search_food.setPointSize(20)
        self.lineEdit.setFont(font_search_food)
        self.lineEdit.setFixedHeight(50)
        self.lineEdit.setFixedWidth(800)
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit.returnPressed.connect(lambda: (self.save_input(), self.save_current_data_to_excel()))  # Enter 키를 눌렀을 때도 검색 후 저장
        # self.lineEdit.returnPressed.connect(self.search_in_order_table) 
        self.searchButton = QtWidgets.QPushButton(self.page_3)
        self.searchButton.setText("검색")
        self.searchButton.setFixedHeight(80)
        self.searchButton.setObjectName("searchButton")
        self.searchButton.clicked.connect(lambda: (self.save_input(), self.save_current_data_to_excel()))  # 버튼을 눌렀을 때도 검색 후 저장


        self.label_4_layout = QHBoxLayout()
        self.label_4_layout.addItem(QSpacerItem(100, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        self.label_4_layout.addWidget(self.label_4, alignment=QtCore.Qt.AlignCenter)
        self.label_4_layout.addItem(QSpacerItem(100, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))

        self.date_label = QtWidgets.QLabel(self.page_3)
        self.date_label.setObjectName("date_label")
        current_date = QDate.currentDate().toString('yyyy-MM-dd')
        self.date_label.setText(current_date)
        self.label_4_layout.addWidget(self.date_label, alignment=QtCore.Qt.AlignRight)

        self.page3Layout = QVBoxLayout(self.page_3)
        self.page3Layout.addWidget(self.homeButton3, alignment=QtCore.Qt.AlignLeft)
        self.page3Layout.addLayout(self.label_4_layout, stretch=0)

        self.searchLayout = QHBoxLayout()
        self.searchLayout.addWidget(self.lineEdit)
        self.searchLayout.addWidget(self.searchButton)

        self.page3Layout.addStretch()
        self.page3Layout.addLayout(self.searchLayout)

        self.table_widget = QTableWidget(self.page_3)
        self.table_widget.setFixedHeight(400)
        # self.table_widget.itemChanged.connect(self.on_item_changed)  # 셀 내용 변경 이벤트 연결
        self.page3Layout.addWidget(self.table_widget)

        # QTableWidget 아래에 빈 공간과 '추가', '삭제' 버튼 추가
        self.buttonLayout = QHBoxLayout()
        self.buttonLayout.addStretch()  # 왼쪽에 빈 공간 추가
        self.addButton = QPushButton("추가")
        self.deleteButton = QPushButton("삭제")
        self.buttonLayout.addWidget(self.addButton)
        self.buttonLayout.addWidget(self.deleteButton)
        self.page3Layout.addLayout(self.buttonLayout)

        self.page3Layout.addStretch(1)
        self.page3Layout.setSpacing(10)

        StackedWidget.addWidget(self.page_3)

        # 삭제 버튼에 대한 이벤트 연결
        self.deleteButton.clicked.connect(lambda: (self.delete_selected_row(), self.save_current_data_to_excel()))

        # 추가 버튼에 대한 이벤트 연결
        self.addButton.clicked.connect(lambda: (self.add_new_row(), self.save_current_data_to_excel()))

    
    # 페이지 4 - 원가율 페이지 설정
    def setup_cost_rate_page(self, StackedWidget):

        # self.page_4 = QtWidgets.QWidget()
        # self.page_4.setObjectName("page_4")

        # self.homeButton4 = QtWidgets.QPushButton(self.page_4)
        # self.homeButton4.setText("홈(사진)")
        # self.homeButton4.setAutoDefault(False)
        # self.homeButton4.clicked.connect(lambda: (StackedWidget.setCurrentIndex(0), self.save_current_data_to_excel()))

        # font = QtGui.QFont()
        # font.setPointSize(25)
        # self.label_5 = QtWidgets.QLabel(self.page_4)
        # self.label_5.setFont(font)
        # self.label_5.setObjectName("label_5")
        # self.label_5.setText("      원가율 페이지")

        # self.page4Layout = QVBoxLayout(self.page_4)
        # self.page4Layout.addWidget(self.homeButton4, alignment=QtCore.Qt.AlignLeft)
        # self.page4Layout.addWidget(self.label_5, alignment=QtCore.Qt.AlignCenter)

        # StackedWidget.addWidget(self.page_4)

        # 페이지 4 (원가율)
        self.page_4 = QtWidgets.QWidget()
        self.page_4.setObjectName("page_4")

        self.homeButton4 = QtWidgets.QPushButton(self.page_4)
        self.homeButton4.setText("홈(사진)")
        self.homeButton4.setAutoDefault(False)
        self.homeButton4.clicked.connect(lambda: StackedWidget.setCurrentIndex(0))

        self.label_5 = QtWidgets.QLabel(self.page_4)
        font = QtGui.QFont()
        font.setPointSize(40)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_5.setText("원가율 페이지")

        self.page4Layout = QVBoxLayout(self.page_4)
        self.page4Layout.addWidget(self.homeButton4, alignment=QtCore.Qt.AlignLeft)
        self.page4Layout.addWidget(self.label_5, alignment=QtCore.Qt.AlignCenter)

        # ReadOnlyExcelViewer 추가
        self.readOnlyExcelViewer = ReadOnlyExcelViewer(data_path_cost_ratio, '원가율')
        self.page4Layout.addWidget(self.readOnlyExcelViewer)

        StackedWidget.addWidget(self.page_4)
    #페이지ㅡ 5 - 발주서 페이지 설정
    def setup_order_form_page(self, StackedWidget):
        self.page_5 = QtWidgets.QWidget()
        self.page_5.setObjectName("page_5")

        self.homeButton5 = QtWidgets.QPushButton(self.page_5)
        self.homeButton5.setText("홈(사진)")
        self.homeButton5.setAutoDefault(False)
        self.homeButton5.clicked.connect(lambda: (StackedWidget.setCurrentIndex(0), ))
        #self.save_current_data_to_excel()))

        font = QtGui.QFont()
        font.setPointSize(25)
        self.label_6 = QtWidgets.QLabel(self.page_5)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.label_6.setText("      발주서 페이지")

        # 검색, lineEdit 레이아웃 추가
        self.search_layout = QHBoxLayout()
        self.lineEdit = QLineEdit(self.page_5)
        self.lineEdit.setFixedSize(300, 40)  # 가로 300, 세로 40으로 설정
        self.lineEdit.setPlaceholderText("찾을 음식을 입력")
        self.search_button = QPushButton("검색", self.page_5)
        self.search_button.setFixedSize(100, 40)  # 가로 100, 세로 40으로 설정

        # placeholder 텍스트 크기 설정
        self.lineEdit.setStyleSheet("font-size: 20px;")  

        # 엔터 키 이벤트 연결
        self.lineEdit.returnPressed.connect(self.search_in_order_table)  

        # 마우스 이벤트 얀결
        self.search_button.clicked.connect(self.search_in_order_table)
        self.search_layout.addWidget(self.lineEdit)
        self.search_layout.addWidget(self.search_button)
        
        # 탑 레이아웃에 위젯 추가하고 위치 선정
        self.topLayout = QHBoxLayout()
        self.topLayout.addWidget(self.homeButton5, alignment=QtCore.Qt.AlignLeft)
        self.topLayout.addWidget(self.label_6, alignment=QtCore.Qt.AlignCenter)

        # 넣을까 뺄까
        self.topLayout.addItem(QSpacerItem(100, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        
        self.topLayout.addLayout(self.search_layout)

        # 좌측에 테이블 추가
        self.fixed_amount_table = QTableWidget(self.page_5)
        self.fixed_amount_table.setObjectName("fixed_amount_table")
        # 좌측 테이블 크기 설정
        self.fixed_amount_table.setFixedWidth(self.page_5.size().width() - 50)

        # 우측에 테이블 추가
        self.order_table = QTableWidget(self.page_5)
        self.order_table.setObjectName("order_table")

        # 우측 테이블 크기 설정
        self.order_table.setFixedWidth(self.page_5.size().width()- 50)

        # 우측 테이블 행 더블 클릭 시 좌측 테이블에 추가하는 이벤트 연결
        self.order_table.itemDoubleClicked.connect(self.add_row_to_fixed_amount_table)

        # 엑셀로 내보내기 버튼
        self.export_button = QtWidgets.QPushButton(self.page_5)
        self.export_button.setText("엑셀로 내보내기")
        self.export_button.clicked.connect(self.export_to_excel)

        self.page5Layout = QVBoxLayout(self.page_5)
        self.page5Layout.addLayout(self.topLayout, stretch=0)  # 상단 레이아웃을 최상단에 배치

        # 좌측에 스페이서를 추가하고 좌우로 테이블 배치
        self.page5ContentLayout = QHBoxLayout()
        self.page5ContentLayout.addWidget(self.fixed_amount_table)
        self.page5ContentLayout.addItem(QSpacerItem(10, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        self.page5ContentLayout.addWidget(self.order_table)

        self.page5Layout.addLayout(self.page5ContentLayout)

        # 좌측 하단에 삭제 버튼 추가
        self.delete_button = QPushButton("삭제")
        self.delete_button.clicked.connect(self.delete_selected_row_for_page5)



        # 우측 하단에 엑셀로 내보내기 버튼 추가
        self.bottomLayout = QHBoxLayout()
        self.bottomLayout.addStretch()
        self.bottomLayout.addWidget(self.export_button, alignment=QtCore.Qt.AlignRight)
        self.bottomLayout.addWidget(self.delete_button, alignment=QtCore.Qt.AlignLeft)
        self.page5Layout.addLayout(self.bottomLayout)

        StackedWidget.addWidget(self.page_5)
        
        # 발주 페이지 초기화 시 엑셀 파일 로드
        self.load_order_table()
        self.load_fixed_amount_table()

    # 우측 테이블의 상단 검색 함수
    def search_in_order_table(self):
        search_text = self.lineEdit.text().lower()
        for row in range(self.order_table.rowCount()):
            match = False
            for col in range(self.order_table.columnCount()):
                item = self.order_table.item(row, col)
                if item and search_text in item.text().lower():
                    match = True
                    break
            self.order_table.setRowHidden(row, not match)

    # 더블 클릭 시 우측 테이블의 행 -> 좌측 테이블의 행으로 column값 매핑해서 삽입
    def add_row_to_fixed_amount_table(self, item):
        row = item.row()
        # 좌측 테이블에 새 행 추가
        new_row_index = self.fixed_amount_table.rowCount()
        self.fixed_amount_table.insertRow(new_row_index)

        # 좌측 테이블에 필요한 데이터 매핑
        right_to_left_col_mapping = {
            7: 0,  # 업체명 -> 발주업체
            1: 1,  # 품목 -> 품목
            3: 3,  # 단가 -> 단가
            8: 6,  # 비고 -> 비고
        }

        for right_col, left_col in right_to_left_col_mapping.items():
            order_item = self.order_table.item(row, right_col)
            new_item = QTableWidgetItem(order_item.text() if order_item else "")
            self.fixed_amount_table.setItem(new_row_index, left_col, new_item)

        # 수량과 예상금액 설정 (수량은 직접 입력할 예정이므로 빈 셀로 둠)
        quantity_item = QTableWidgetItem("")
        quantity_item.setFlags(quantity_item.flags() | Qt.ItemIsEditable)
        self.fixed_amount_table.setItem(new_row_index, 2, quantity_item)

        # 예상금액 = 수량 * 단가, 단가 가져오기
        unit_price_item = self.order_table.item(row, 3)
        if unit_price_item:
            try:
                unit_price = float(unit_price_item.text())
            except ValueError:
                unit_price = 0.0

            # 예상금액 계산 함수
            def calculate_expected_amount():
                # 수량에 추가된 입력 값 가져옴(행 index, 수량 index)
                quantity_text = self.fixed_amount_table.item(new_row_index, 2).text()
                # 정수 추출 ex) 2 박스, 3 망, 6 봉지
                quantity_numbers = re.findall(r'\d+', quantity_text)
                if quantity_numbers:
                    # 2 박스 -> 2
                    quantity = float(quantity_numbers[0])
                    expected_amount = quantity * unit_price
                    self.fixed_amount_table.itemChanged.disconnect(calculate_expected_amount)
                    self.fixed_amount_table.setItem(new_row_index, 4, QTableWidgetItem(str(expected_amount)))
                    self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)
                else:
                    self.fixed_amount_table.itemChanged.disconnect(calculate_expected_amount)
                    self.fixed_amount_table.setItem(new_row_index, 4, QTableWidgetItem(""))
                    self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)

            # 수량 셀에 대한 이벤트 연결 (수량이 입력되면 예상금액 자동 계산)
            self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)

    # 좌측 테이블에 들어갈 엑셀 파일 로드, 컬럼 값 세팅
    def load_fixed_amount_table(self):
        file_name = './고정금액.xlsx'  # 자동으로 불러올 파일 이름
        try:
            df = pd.read_excel(file_name)
            if df.empty:
                print('hello')
                return
            
            self.fixed_amount_table.setRowCount(df.shape[0])
            self.fixed_amount_table.setColumnCount(df.shape[1])
            self.fixed_amount_table.setHorizontalHeaderLabels(df.columns)
            
            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    self.fixed_amount_table.setItem(row, col, QTableWidgetItem(str(df.iat[row, col])))
        except Exception as e:
            QMessageBox.critical(None, 'Error', str(e))

    # 우측 테이블에 들어갈 엑셀 파일 로드, 컬럼 값 세팅
    def load_order_table(self):
        file_name = './매입.xlsx'  # 자동으로 불러올 파일 이름
        try:
            df = pd.read_excel(file_name)
            if df.empty:
                return
            
            self.order_table.setRowCount(df.shape[0])
            self.order_table.setColumnCount(df.shape[1])
            self.order_table.setHorizontalHeaderLabels(df.columns)
            
            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    self.order_table.setItem(row, col, QTableWidgetItem(str(df.iat[row, col])))
        except Exception as e:
            QMessageBox.critical(None, 'Error', str(e))
    # 좌측 테이블을 발주서 엑셀 파일로 작성
    def export_to_excel(self):
        rows = self.fixed_amount_table.rowCount()
        cols = self.fixed_amount_table.columnCount()
        data = []

        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.fixed_amount_table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        df = pd.DataFrame(data, columns=[self.fixed_amount_table.horizontalHeaderItem(i).text() for i in range(cols)])
        
        # 현재 날짜를 'yymmdd' 형식으로 가져오기
        current_date = datetime.now().strftime("%y%m%d")

        # 파일 이름에 날짜 추가
        file_name = f'./{current_date}_발주서.xlsx'
        df.to_excel(file_name, index=False)

        # 엑셀 파일 불러오기 및 병합 작업 수행
        wb = load_workbook(file_name)
        ws = wb.active

        # 15, 16, 17번째 row의 '발주업체' column을 병합 (1-based index)
        ws.merge_cells(start_row=15, start_column=1, end_row=17, end_column=1)
        
        # 병합된 셀 가운데 정렬
        merged_cell = ws.cell(row=15, column=1)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 엑셀 파일 저장
        wb.save(file_name)

        QMessageBox.information(None, '완료', '엑셀 파일로 저장되었습니다.')

    # 발주서 페이지의 좌측 테이블 삭제 버튼 메소드
    def delete_selected_row_for_page5(self):
        try: # 무한 루프 방지
            self.fixed_amount_table.itemChanged.disconnect()
        except Exception as e:
            print(f"Error: {e}")
        selected_row = self.fixed_amount_table.currentRow()
        if selected_row > 0:
            reply = QMessageBox.question(self.fixed_amount_table, '삭제 확인', '정말로 삭제하시겠습니까?',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                self.fixed_amount_table.removeRow(selected_row)
                # self.df.drop(self.df.index[selected_row], inplace=True)  # 데이터프레임에서도 삭제


    def retranslateUi(self, StackedWidget):
        _translate = QtCore.QCoreApplication.translate
        StackedWidget.setWindowTitle(_translate("StackedWidget", "BapuriFNB"))

    def load_excel(self):
        file_name = './매입.xlsx'
        try:
            self.excel_data = pd.read_excel(file_name, sheet_name=None)  # 모든 시트 로드
            self.df = self.excel_data['물품목록list']  # 필요한 시트만 작업
            self.display_data(self.df)
        except Exception as e:
            QMessageBox.critical(None, 'Error', str(e))
    # 데이터를 테이블 위젯에 표시
    def display_data(self, df):

        self.table_widget.setRowCount(df.shape[0])
        self.table_widget.setColumnCount(df.shape[1])
        self.table_widget.setHorizontalHeaderLabels(df.columns)

        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                self.table_widget.setItem(row, col, QTableWidgetItem(str(df.iat[row, col])))
    # 입력받은 데이터로 검색 후 표시
    def save_input(self):
        user_input = self.lineEdit.text()
        if user_input and self.df is not None:
            filtered_df = self.df[self.df.apply(lambda row: row.astype(str).str.contains(user_input).any(), axis=1)]
            self.display_data(filtered_df)

    def delete_selected_row(self):
        # 선택된 행 삭제 후 엑셀 파일 업데이트
        current_row = self.table_widget.currentRow()
        if current_row >= 0:
            reply = QMessageBox.question(self.table_widget, '삭제 확인', '정말로 삭제하시겠습니까?',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                self.table_widget.removeRow(current_row)
                self.df.drop(self.df.index[current_row], inplace=True)  # 데이터프레임에서도 삭제

    def add_new_row(self):
        # 새로운 행 추가 후 엑셀 파일 업데이트
        row_position = self.table_widget.rowCount()
        self.table_widget.insertRow(row_position)
        for i in range(self.table_widget.columnCount()):
            self.table_widget.setItem(row_position, i, QTableWidgetItem("새 데이터"))

    def save_current_data_to_excel(self):
        # 현재 데이터를 엑셀 파일에 저장
        file_name = './매입.xlsx'
        rows = self.table_widget.rowCount()
        cols = self.table_widget.columnCount()
        data = []

        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.table_widget.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)

        df = pd.DataFrame(data, columns=[self.table_widget.horizontalHeaderItem(i).text() for i in range(cols)])
        self.excel_data['물품목록list'] = df  # 수정된 데이터프레임을 원래의 시트에 반영
        with pd.ExcelWriter(file_name) as writer:
            for sheet_name, sheet_df in self.excel_data.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)









class MyStackedWidget(QtWidgets.QStackedWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_StackedWidget()
        self.ui.setupUi(self)

    def closeEvent(self, event):
        # 앱 종료 확인
        reply = QtWidgets.QMessageBox.question(self, '종료 확인', '종료 하시겠습니까?',
                                               QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                                               QtWidgets.QMessageBox.Yes)
        if reply == QtWidgets.QMessageBox.Yes:
            #self.ui.save_current_data_to_excel()  # 종료 시 데이터 저장
            event.accept()
        else:
            event.ignore()

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    StackedWidget = MyStackedWidget()
    StackedWidget.show()
    sys.exit(app.exec_())
