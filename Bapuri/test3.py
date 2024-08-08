# 시장조사 수정중
from PyQt5 import QtCore, QtGui, QtWidgets
from datetime import datetime
import pandas as pd
import sys, re, os
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox, QVBoxLayout, QHBoxLayout, QSplitter, \
    QLabel, QTableWidget, QTableWidgetItem, QToolButton, QLineEdit, QSpacerItem, QSizePolicy
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QFont
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QColor

# 엑셀 파일 경로 설정
data_path_purchase = '2024_식자재매입_test.xlsx'  # 매출매입 페이지: 매입표
data_path_sales = '2024_식자재매출_test5.xlsx'  # 매출매입 페이지: 매출표
data_path_cost_ratio = '2024_식자재매출_test5.xlsx'  # 원가율 페이지: 원가율표
data_path_total_purchase = 'combined_totals.xlsx'  # 매출매입 페이지: 매입 총합 데이터
data_path_company_total_purchase = 'company_total_purchase.xlsx'

class ExcelTreeView(QtWidgets.QWidget):
    def __init__(self, filePath, label='매출'):
        super().__init__()
        self.filePath = filePath
        self.label = label
        self.initUI()
        self.loadExcelData(self.filePath)

    def initUI(self):
        self.layout = QtWidgets.QVBoxLayout(self)

        # 매출 레이블 추가
        self.labelWidget = QLabel(self.label, self)
        self.labelWidget.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.labelWidget)

        self.treeView = QtWidgets.QTreeView(self)
        self.layout.addWidget(self.treeView)

        self.model = QtGui.QStandardItemModel()
        self.treeView.setModel(self.model)

        # 열 크기 조정 모드를 설정
        header = self.treeView.header()
        header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)

        buttonLayout = QHBoxLayout()

        self.btnSave = QPushButton('Save Changes', self)
        self.btnSave.clicked.connect(self.saveChanges)
        buttonLayout.addWidget(self.btnSave)

        self.layout.addLayout(buttonLayout)
        self.setLayout(self.layout)

    def loadExcelData(self, filePath):
        try:
            # Load the company data
            df_company = pd.read_excel(data_path_company_total_purchase)
            df_company['일자'] = pd.to_datetime(df_company['일자'], format='%Y-%m-%d', errors='coerce').dt.date
            df_company = df_company.dropna(subset=['일자'])

            # Load the main data
            self.df = pd.read_excel(filePath)

            monthly_data = self.df.iloc[0:12].copy()
            daily_data = self.df.iloc[12:].copy()

            # Convert '일자' column to datetime format, drop rows where conversion fails
            daily_data['일자'] = pd.to_datetime(daily_data['일자'], format='%Y-%m-%d', errors='coerce').dt.date
            daily_data = daily_data.dropna(subset=['일자'])

            # Extract month from '일자'
            daily_data['월'] = pd.to_datetime(daily_data['일자']).dt.month

            self.model.clear()  # Clear the existing model data

            # '월' 열을 제외한 나머지 열 이름 추출
            header_labels = [col for col in daily_data.columns if col != '월']
            # 머리글 설정
            self.model.setHorizontalHeaderLabels(header_labels)

            current_date = pd.to_datetime('today').date()

            for month, day_group in daily_data.groupby('월'):
                month_item = [QtGui.QStandardItem(str(cell)) for cell in monthly_data.iloc[int(month) - 1].fillna(0)]
                self.model.appendRow(month_item)
                month_item_ptr = self.model.indexFromItem(month_item[0])  # Get the QModelIndex for the month item

                for day in day_group.itertuples(index=False):
                    # 마지막 열을 제외한 나머지 셀 값으로 QStandardItem 객체 생성
                    day_item = []
                    for idx, cell in enumerate(day[:-1]):
                        item = QtGui.QStandardItem(str(cell))
                        if header_labels[idx] == '금액':
                            item.setBackground(QtGui.QColor('skyblue'))
                        elif header_labels[idx] != '비고':
                            item.setEditable(False)
                        day_item.append(item)

                    month_item[0].appendRow(day_item)
                    day_item_ptr = month_item[0].index().child(month_item[0].rowCount() - 1, 0)  # Get the QModelIndex for the day item

                    # Expand the item if it matches the current date
                    if day.일자 == current_date:
                        self.treeView.expand(month_item_ptr)
                        self.treeView.expand(day_item_ptr)
                        self.treeView.scrollTo(day_item_ptr)
                        self.treeView.setCurrentIndex(day_item_ptr)

                    company_day_data = df_company[df_company['일자'] == day.일자]
                    for company in company_day_data.itertuples(index=False):
                        company = company[1:2] + ('',) + company[3:]
                        row_items = [QtGui.QStandardItem(str(cell)) for cell in company]
                        day_item[0].appendRow(row_items)
        except Exception as e:
            print(f"Error loading Excel data: {e}")

    def saveChanges(self):
        # Load data from model to DataFrame
        all_data = []

        # Append parent data (months)
        for i in range(12):
            parent_item = [self.model.item(i, col).text() if self.model.item(i, col) else '' for col in range(self.model.columnCount())]
            all_data.append(parent_item)

        # Append children data (daily) and calculate monthly total
        monthly_totals = [0] * 12
        for i in range(12):
            parent_item = self.model.item(i)
            if parent_item:
                for j in range(parent_item.rowCount()):
                    child_items = [parent_item.child(j, col).text() if parent_item.child(j, col) else '' for col in range(parent_item.columnCount())]
                    all_data.append(child_items)
                    # Add to monthly total
                    monthly_totals[i] += float(child_items[1]) if child_items[1] else 0

        # Update parent items with monthly totals
        for i in range(12):
            all_data[i][1] = monthly_totals[i]

        # Create DataFrame
        df = pd.DataFrame(all_data, columns=self.df.columns)
        df['금액'] = pd.to_numeric(df['금액'], errors='coerce').fillna(0)

        # Load purchase total data
        sales_df = pd.read_excel(self.filePath)

        sales_df['매입액'] = pd.to_numeric(sales_df['매입액'], errors='coerce').fillna(0)

        # Calculate cost ratio
        df['원가율'] = ((sales_df['매입액'] / df['금액'] * 100)).round(1).fillna(0)
        df['매입액'] = sales_df['매입액']

        # Save to Excel
        df.to_excel(self.filePath, index=False)


        # 매출 데이터와 매입 데이터 병합
        company_purchase_df = pd.read_excel(data_path_company_total_purchase)
        df = df.iloc[12:]
        df['일자'] = pd.to_datetime(df['일자'], format='%Y-%m-%d', errors='coerce')
        if '금액' in company_purchase_df.columns:
            company_purchase_df = company_purchase_df.drop(columns=['금액'])


        company_purchase_df = pd.merge(company_purchase_df, df[['일자', '금액']], on='일자', how='left')

        # '금액' 열을 3번째 위치로 이동
        cols = list(company_purchase_df.columns)
        cols.insert(2, cols.pop(cols.index('금액')))
        company_purchase_df = company_purchase_df[cols]

        company_purchase_df['총금액'] = pd.to_numeric(company_purchase_df['총금액'], errors='coerce').fillna(0)

        # 원가율 계산 (매입액 / 매출액)
        company_purchase_df['원가율'] = (company_purchase_df['총금액'] / company_purchase_df['금액'] * 100).round(1)

        # 매출값 열 제거
        # company_purchase_df = company_purchase_df.drop(columns=['금액'])

        company_purchase_df.to_excel(data_path_company_total_purchase, index=False)

        # Reload the updated data
        self.loadExcelData(self.filePath)

        QMessageBox.information(self, 'Success', 'Changes saved successfully!')

class ExcelViewer(QWidget):
    def __init__(self, filePath, label, rightViewer):
        super().__init__()
        self.filePath = filePath
        self.label = label
        self.rightViewer = rightViewer
        self.initUI()
        self.loadExcelData(self.filePath)

    def initUI(self):
        self.layout = QVBoxLayout()

        self.labelWidget = QLabel(self.label, self)
        self.labelWidget.setAlignment(Qt.AlignCenter)

        self.tableWidget = QTableWidget(self)
        self.layout.addWidget(self.labelWidget)
        self.layout.addWidget(self.tableWidget)

        buttonLayout = QHBoxLayout()

        self.btnAddRow = QPushButton('Add Row', self)
        self.btnAddRow.clicked.connect(self.addRow)
        buttonLayout.addWidget(self.btnAddRow)

        self.btnRemoveRow = QPushButton('Remove Selected Row', self)
        self.btnRemoveRow.clicked.connect(self.removeRow)
        buttonLayout.addWidget(self.btnRemoveRow)

        self.btnSave = QPushButton('Save Changes', self)
        self.btnSave.clicked.connect(self.saveChanges)
        buttonLayout.addWidget(self.btnSave)

        self.layout.addLayout(buttonLayout)
        self.setLayout(self.layout)

    def loadExcelData(self, filePath):
        self.df = pd.read_excel(filePath)

        if self.df is not None:
            self.df['일자'] = pd.to_datetime(self.df['일자']).dt.strftime('%Y-%m-%d')

            # Convert specific columns to integer if they are numeric
            for column in self.df.columns:
                if pd.api.types.is_numeric_dtype(self.df[column]):
                    self.df[column] = self.df[column].fillna(0).astype(int)

            self.tableWidget.setRowCount(self.df.shape[0])
            self.tableWidget.setColumnCount(self.df.shape[1])
            self.tableWidget.setHorizontalHeaderLabels(self.df.columns)

            for i in range(self.df.shape[0]):
                for j in range(self.df.shape[1]):
                    item = QTableWidgetItem(str(self.df.iat[i, j]))
                    if self.df.columns[j] in ['공급가액', '부가세']:
                        item.setBackground(QColor(245, 245, 245))  # 옅은 회색
                    self.tableWidget.setItem(i, j, item)

            # Scroll to the bottom
            self.tableWidget.scrollToBottom()
            # Store the column widths
            self.storeColumnWidths()
            # Resize columns to contents
            self.resizeColumns()

    def storeColumnWidths(self):
        self.columnWidths = [self.tableWidget.columnWidth(col) for col in range(self.tableWidget.columnCount())]

    def restoreColumnWidths(self):
        for col, width in enumerate(self.columnWidths):
            self.tableWidget.setColumnWidth(col, width)

    def resizeColumns(self):
        # 특정 열의 너비를 고정값으로 설정 ('품목' 열을 예시로 사용)
        column_index = self.df.columns.get_loc('품목') if '품목' in self.df.columns else None
        if column_index is not None:
            self.tableWidget.setColumnWidth(column_index, 150)  # '품목' 열의 너비를 150으로 고정

        # 나머지 열의 크기를 내용에 맞게 자동 조정
        for col in range(self.tableWidget.columnCount()):
            if col != column_index:
                self.tableWidget.resizeColumnToContents(col)

    def addRow(self):
        self.storeColumnWidths()  # Store current column widths before adding a row
        currentRow = self.tableWidget.currentRow()
        self.tableWidget.insertRow(currentRow + 1)
        # Add empty QTableWidgetItem objects to the new row
        for col in range(self.tableWidget.columnCount()):
            self.tableWidget.setItem(currentRow + 1, col, QTableWidgetItem(""))
        # Restore column widths after adding a row
        self.restoreColumnWidths()

    def removeRow(self):
        self.storeColumnWidths()  # Store current column widths before removing a row
        currentRow = self.tableWidget.currentRow()
        if currentRow >= 0:
            self.tableWidget.removeRow(currentRow)
            # Restore column widths after removing a row
            self.restoreColumnWidths()
        else:
            QMessageBox.warning(self, 'Warning', 'Please select a row to remove.')

    def saveChanges(self):
        new_data = []
        for i in range(self.tableWidget.rowCount()):
            row_data = []
            for j in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(i, j)
                if item:
                    row_data.append(item.text())
                else:
                    row_data.append(None)
            new_data.append(row_data)

        new_df = pd.DataFrame(new_data, columns=self.df.columns)

        for column in self.df.columns:
            if pd.api.types.is_numeric_dtype(self.df[column]):
                new_df[column] = pd.to_numeric(new_df[column], errors='coerce')
            else:
                new_df[column] = new_df[column].astype(self.df[column].dtype)

        self.df = new_df

        self.df['공급가액'] = self.df['단가'] * self.df['수량']
        self.df['부가세'] = self.df.apply(lambda row: row['공급가액'] / 10 if row['부가세'] != 0 else 0, axis=1) # 부가세 계산: 부가세 값이 0이 아닌 경우 공급가액의 10%, 0인 경우 그대로 0

        self.df.to_excel(self.filePath, index=False)

        self.loadExcelData(self.filePath)

        # Save monthly totals and daily data to another Excel file
        self.saveMonthlyAndDailyTotals()

        # Save daily totals grouped by date and supplier
        self.saveGroupedByDateAndSupplier()

        # 매출부분(트리뷰)의 저장 함수 호출
        self.rightViewer.saveChanges()

        #QMessageBox.information(self, 'Success', 'Changes saved successfully!')

    def saveMonthlyAndDailyTotals(self):
        # Ensure '일자' column is in datetime format
        self.df['일자'] = pd.to_datetime(self.df['일자'])

        # Create lists to hold the monthly and daily data
        monthly_data = []
        daily_data = []

        # Calculate monthly totals
        for month in range(1, 13):
            monthly_sum = self.df[self.df['일자'].dt.month == month]['금액'].sum()
            monthly_data.append({'일자': f'{month}월', '총금액': monthly_sum if not pd.isna(monthly_sum) else 0})

        # Calculate daily totals
        for date in pd.date_range(start='2024-01-01', end='2024-12-31'):
            daily_sum = self.df[self.df['일자'] == date]['금액'].sum()
            daily_data.append({'일자': date.strftime('%Y-%m-%d'), '총금액': daily_sum if not pd.isna(daily_sum) else 0})

        # Combine monthly and daily data
        combined_data = monthly_data + daily_data

        # Create a DataFrame for combined data
        combined_df = pd.DataFrame(combined_data)

        # Load the existing sales Excel file
        sales_df = pd.read_excel(data_path_sales)

        # Merge combined_df with df_sales on '일자' column
        sales_df = sales_df.merge(combined_df, on='일자', how='left')

        # Assign the '총금액' to '매입액' in df_sales
        sales_df['매입액'] = sales_df['총금액']
        sales_df.drop(columns=['총금액'], inplace=True)

        # Save combined totals to a new Excel file
        sales_df.to_excel(data_path_sales, index=False)
        combined_df.to_excel(data_path_total_purchase, index=False)

    def saveGroupedByDateAndSupplier(self):
        # Ensure '일자' column is in datetime format
        #self.df['일자'] = pd.to_datetime(self.df['일자'])
        self.df['일자'] = pd.to_datetime(self.df['일자'], format='%Y-%m-%d', errors='coerce').dt.date

        # Group by '일자' and '업체명' and sum the '금액'
        grouped_df = self.df.groupby(['일자', '업체명'], as_index=False)['금액'].sum()
        grouped_df.rename(columns={'금액': '총금액'}, inplace=True)

        # Sort by '일자'
        grouped_df.sort_values(by='일자', inplace=True)

        # Save to a new Excel file
        grouped_df.to_excel(data_path_company_total_purchase, index=False)

class ReadOnlyExcelViewer(QWidget):
    def __init__(self, filePath, label):
        super().__init__()
        self.filePath = filePath
        self.label = label
        self.initUI()
        self.loadExcelData(self.filePath)

    def initUI(self):
        self.layout = QVBoxLayout()

        self.labelWidget = QLabel(self.label, self)
        self.labelWidget.setAlignment(Qt.AlignCenter)

        self.tableWidget = QTableWidget(self)
        self.layout.addWidget(self.labelWidget)
        self.layout.addWidget(self.tableWidget)

        buttonLayout = QHBoxLayout()

        self.btnShowMonthly = QPushButton('Show Monthly Data', self)
        self.btnShowMonthly.clicked.connect(self.showMonthlyData)
        buttonLayout.addWidget(self.btnShowMonthly)

        self.btnShowDaily = QPushButton('Show Daily Data', self)
        self.btnShowDaily.clicked.connect(self.showDailyData)
        buttonLayout.addWidget(self.btnShowDaily)

        self.layout.addLayout(buttonLayout)
        self.setLayout(self.layout)

    def loadExcelData(self, filePath):
        self.df = pd.read_excel(filePath)
        self.df.iloc[:12, self.df.columns.get_loc('일자')] = self.df.iloc[:12, self.df.columns.get_loc('일자')].astype(str)
        self.df.iloc[12:, self.df.columns.get_loc('일자')] = pd.to_datetime(self.df.iloc[12:, self.df.columns.get_loc('일자')], errors='coerce').dt.strftime('%Y-%m-%d')
        self.showMonthlyData()  # 기본적으로 월별 데이터를 표시

    def showMonthlyData(self):
        monthly_df = self.df.iloc[:12]  # 첫 12행이 월별 데이터
        self.displayData(monthly_df)

    def showDailyData(self):
        daily_df = self.df.iloc[12:]  # 13행부터가 일별 데이터
        self.displayData(daily_df)

    def displayData(self, data):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(data.shape[0])
        self.tableWidget.setColumnCount(data.shape[1])
        self.tableWidget.setHorizontalHeaderLabels(data.columns)

        for i in range(data.shape[0]):
            for j in range(data.shape[1]):
                item = QTableWidgetItem(str(data.iat[i, j]))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)  # 읽기 전용으로 설정
                self.tableWidget.setItem(i, j, item)

        # Scroll to the bottom
        self.tableWidget.scrollToBottom()

    def saveChanges(self):
        pass

    def addRow(self):
        pass

    def removeRow(self):
        pass

class OrderFormViewer(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.load_order_table()
        self.load_fixed_amount_table()

    def initUI(self):
        self.layout = QVBoxLayout()

        self.homeButton = ClickableLabel(self)
        home_pixmap = QPixmap("./home.png")  # 이미지 파일 경로 설정
        home_pixmap = home_pixmap.scaled(50, 50, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)  # 적절한 크기로 조정
        self.homeButton.setPixmap(home_pixmap)
        self.homeButton.setAlignment(QtCore.Qt.AlignLeft)

        self.label = QLabel("발주서 페이지", self)
        font = QtGui.QFont()
        font.setPointSize(25)
        self.label.setFont(font)

        self.search_layout = QHBoxLayout()
        self.lineEdit = QLineEdit(self)
        self.lineEdit.setFixedSize(300, 40)
        self.lineEdit.setPlaceholderText("찾을 음식을 입력")
        self.search_button = QPushButton("검색", self)
        self.search_button.setFixedSize(100, 40)

        self.lineEdit.setStyleSheet("font-size: 20px;")
        self.lineEdit.returnPressed.connect(self.search_in_order_table)
        self.search_button.clicked.connect(self.search_in_order_table)

        self.search_layout.addWidget(self.lineEdit)
        self.search_layout.addWidget(self.search_button)

        self.topLayout = QHBoxLayout()
        self.topLayout.addWidget(self.homeButton, alignment=QtCore.Qt.AlignLeft)
        self.topLayout.addWidget(self.label, alignment=QtCore.Qt.AlignCenter)
        self.topLayout.addItem(QSpacerItem(100, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        self.topLayout.addLayout(self.search_layout)

        self.fixed_amount_table = QTableWidget(self)
        self.fixed_amount_table.setFixedWidth(self.size().width() - 50)

        self.order_table = QTableWidget(self)
        self.order_table.setFixedWidth(self.size().width() - 50)
        self.order_table.itemDoubleClicked.connect(self.add_row_to_fixed_amount_table)

        self.export_button = QPushButton("엑셀로 내보내기", self)
        self.export_button.clicked.connect(self.export_to_excel)

        self.contentLayout = QHBoxLayout()
        self.contentLayout.addWidget(self.fixed_amount_table)
        self.contentLayout.addItem(QSpacerItem(10, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        self.contentLayout.addWidget(self.order_table)

        self.bottomLayout = QHBoxLayout()
        self.add_button = QPushButton("추가", self)
        self.add_button.clicked.connect(self.add_row_button)
        self.delete_button = QPushButton("삭제", self)
        self.delete_button.clicked.connect(self.delete_selected_row)

        self.bottomLayout.addWidget(self.add_button, alignment=QtCore.Qt.AlignLeft)
        self.bottomLayout.addWidget(self.delete_button, alignment=QtCore.Qt.AlignLeft)
        self.bottomLayout.addStretch()
        self.bottomLayout.addWidget(self.export_button, alignment=QtCore.Qt.AlignRight)

        self.layout.addLayout(self.topLayout)
        self.layout.addLayout(self.contentLayout)
        self.layout.addLayout(self.bottomLayout)

        self.setLayout(self.layout)

    def load_fixed_amount_table(self):
        file_name = './고정금액.xlsx'  # 자동으로 불러올 파일 이름
        try:
            self.full_df = pd.read_excel(file_name)
            df = pd.read_excel(file_name, skiprows=6)
            if df.empty:
                return
            self.fixed_amount_table.setRowCount(df.shape[0])
            self.fixed_amount_table.setColumnCount(df.shape[1])
            self.fixed_amount_table.setHorizontalHeaderLabels(df.columns)

            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    self.fixed_amount_table.setItem(row, col, QTableWidgetItem(str(df.iat[row, col])))
        except Exception as e:
            QMessageBox.critical(None, 'Error', str(e))

    def load_order_table(self):
        file_name = './2024_식자재매입_test.xlsx'  # 자동으로 불러올 파일 이름
        try:
            df = pd.read_excel(file_name)
            if df.empty:
                return
            self.order_table.setRowCount(df.shape[0])
            self.order_table.setColumnCount(df.shape[1])
            self.order_table.setHorizontalHeaderLabels(df.columns)

            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    item = QTableWidgetItem(str(df.iat[row, col]))
                    item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)  # 셀을 편집 불가능하게 설정
                    self.order_table.setItem(row, col, item)
        except Exception as e:
            QMessageBox.critical(None, 'Error', str(e))

    def add_row_button(self):
        new_row_index = self.fixed_amount_table.rowCount()
        self.fixed_amount_table.insertRow(new_row_index)

        quantity_item = QTableWidgetItem("")
        quantity_item.setFlags(quantity_item.flags() | Qt.ItemIsEditable)
        self.fixed_amount_table.setItem(new_row_index, 2, quantity_item)

        unit_price_item = QTableWidgetItem("")
        unit_price_item.setFlags(unit_price_item.flags() | Qt.ItemIsEditable)
        self.fixed_amount_table.setItem(new_row_index, 3, unit_price_item)

        def calculate_expected_amount():
            quantity_text = self.fixed_amount_table.item(new_row_index, 2).text()
            unit_price_text = self.fixed_amount_table.item(new_row_index, 3).text()
            try:
                quantity = float(re.findall(r'\d+', quantity_text)[0])
                unit_price = float(unit_price_text)
                expected_amount = quantity * unit_price
                self.fixed_amount_table.itemChanged.disconnect(calculate_expected_amount)
                self.fixed_amount_table.setItem(new_row_index, 4, QTableWidgetItem(str(expected_amount)))
                self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)
            except (IndexError, ValueError):
                self.fixed_amount_table.itemChanged.disconnect(calculate_expected_amount)
                self.fixed_amount_table.setItem(new_row_index, 4, QTableWidgetItem(""))
                self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)

        self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)

    def search_in_order_table(self):
        search_text = self.lineEdit.text()
        search_text = search_text.lower().strip()
        if not search_text:
            for row in range(self.order_table.rowCount()):
                self.order_table.setRowHidden(row, False)
            return

        for row in range(self.order_table.rowCount()):
            match = False
            for col in range(self.order_table.columnCount()):
                item = self.order_table.item(row, col)
                if item:
                    item_text = item.text().lower().strip()
                    if search_text in item_text:
                        match = True
                        break
            self.order_table.setRowHidden(row, not match)

    def add_row_to_fixed_amount_table(self, item):
        row = item.row()
        new_row_index = self.fixed_amount_table.rowCount()
        self.fixed_amount_table.insertRow(new_row_index)

        right_to_left_col_mapping = {
            7: 0,  # 업체명 -> 발주업체
            1: 1,  # 품목 -> 품목
            3: 3,  # 단가 -> 단가
            8: 6,  # 비고 -> 비고
        }

        for right_col, left_col in right_to_left_col_mapping.items():
            order_item = self.order_table.item(row, right_col)
            new_item = QTableWidgetItem(order_item.text() if order_item else "")
            self.fixed_amount_table.setItem(new_row_index, left_col, new_item)

        quantity_item = QTableWidgetItem("")
        quantity_item.setFlags(quantity_item.flags() | Qt.ItemIsEditable)
        self.fixed_amount_table.setItem(new_row_index, 2, quantity_item)

        unit_price_item = self.order_table.item(row, 3)
        if unit_price_item:
            try:
                unit_price = float(unit_price_item.text())
            except ValueError:
                unit_price = 0.0

            def calculate_expected_amount():
                quantity_text = self.fixed_amount_table.item(new_row_index, 2).text()
                quantity_numbers = re.findall(r'\d+', quantity_text)
                if quantity_numbers:
                    quantity = float(quantity_numbers[0])
                    expected_amount = quantity * unit_price
                    self.fixed_amount_table.itemChanged.disconnect(calculate_expected_amount)
                    self.fixed_amount_table.setItem(new_row_index, 4, QTableWidgetItem(str(expected_amount)))
                    self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)
                else:
                    self.fixed_amount_table.itemChanged.disconnect(calculate_expected_amount)
                    self.fixed_amount_table.setItem(new_row_index, 4, QTableWidgetItem(""))
                    self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)

            self.fixed_amount_table.itemChanged.connect(calculate_expected_amount)

    def export_to_excel(self):
        rows = self.fixed_amount_table.rowCount()
        cols = self.fixed_amount_table.columnCount()
        data = []

        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.fixed_amount_table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        df = pd.DataFrame(data, columns=[self.fixed_amount_table.horizontalHeaderItem(i).text() for i in range(cols)])

        current_date = datetime.now().strftime("%y%m%d")

        headers = pd.DataFrame([df.columns], columns=df.columns)
        full_data = pd.concat([self.full_df.iloc[:5], headers, df], ignore_index=True, axis=0)

        header_df = full_data.iloc[:21, :]
        data_df = full_data.iloc[21:, :].sort_values(by='발주업체').reset_index(drop=True)
        full_data = pd.concat([header_df, data_df], ignore_index=True)

        first_numeric_index = full_data['예상금액'].apply(pd.to_numeric, errors='coerce').notnull().idxmax()
        full_data.loc[first_numeric_index:, '예상금액'] = pd.to_numeric(full_data.loc[first_numeric_index:, '예상금액'],
                                                                    errors='coerce').fillna(0)
        total_expected_amount = full_data.loc[first_numeric_index:, '예상금액'].sum()

        full_data.loc[first_numeric_index:, '예상금액'] = full_data.loc[first_numeric_index:, '예상금액'].apply(
            lambda x: f"{int(x):,}" if pd.notnull(x) else "")

        if total_expected_amount > 0:
            full_data.loc[first_numeric_index:, '비중율'] = full_data.loc[first_numeric_index:, '예상금액'].str.replace(',',
                                                                                                                 '').astype(
                float) / total_expected_amount * 100
            full_data.loc[first_numeric_index:, '비중율'] = full_data.loc[first_numeric_index:, '비중율'].apply(
                lambda x: f"{x:.2f}%" if pd.notnull(x) else "")
        else:
            full_data.loc[first_numeric_index:, '비중율'] = ""

        total_row = [''] * cols
        total_row[0] = '예상매입'
        total_row[full_data.columns.get_loc('예상금액')] = f"{int(total_expected_amount):,}"
        total_row[full_data.columns.get_loc('비중율')] = '100.00%'
        full_data.loc[len(full_data)] = total_row

        total_sales_row = [''] * cols
        total_sales_row[0] = '예상매출'
        total_sales_row[1] = '예상인원(중/석)'
        full_data.loc[len(full_data) + 1] = total_sales_row

        file_name = f'./{current_date}_발주서.xlsx'
        full_data.to_excel(file_name, index=False)

        wb = load_workbook(file_name)
        ws = wb.active

        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == '예상매입':
                    for target_cell in row:
                        target_cell.fill = yellow_fill

        purple_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in range(8, 14):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.fill = purple_fill
                cell.border = thin_border

        sky_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=7, column=col)
            cell.fill = sky_blue_fill
            cell.border = thin_border

        light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in range(22, 25):
            cell = ws.cell(row=row, column=1)
            cell.fill = light_red_fill
            cell.border = thin_border

        light_red_outer_border = Border(
            left=Side(style='thin', color="FF6666"),
            right=Side(style='thin', color="FF6666"),
            top=Side(style='thin', color="FF6666"),
            bottom=Side(style='thin', color="FF6666")
        )

        for row in range(14, 25):
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                current_border = cell.border

                if row == 14:
                    top = light_red_outer_border.top
                else:
                    top = current_border.top

                if row == 24:
                    bottom = light_red_outer_border.bottom
                else:
                    bottom = current_border.bottom

                if col == 1:
                    left = light_red_outer_border.left
                else:
                    left = current_border.left

                if col == 7:
                    right = light_red_outer_border.right
                else:
                    right = current_border.right

                new_border = Border(left=left, right=right, top=top, bottom=bottom)
                cell.border = new_border

        dark_blue_fill = PatternFill(start_color="6496be", end_color="6496be", fill_type="solid")
        last_row = len(full_data) + 1

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.fill = dark_blue_fill
            cell.border = thin_border

        merge_ranges = ['C3:C5', 'D3:D5', 'E3:E5', 'F3:F5', 'A22:A24']
        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)

        for cell_range in merge_ranges:
            for merged_cell in ws[cell_range]:
                for cell in merged_cell:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(file_name)

        QMessageBox.information(None, '완료', '엑셀 파일로 저장되었습니다.')

    def delete_selected_row(self):
        try:
            self.fixed_amount_table.itemChanged.disconnect()
        except Exception as e:
            print(f"Error: {e}")
        selected_row = self.fixed_amount_table.currentRow()
        if selected_row > 0:
            reply = QMessageBox.question(self.fixed_amount_table, '삭제 확인', '정말로 삭제하시겠습니까?',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                self.fixed_amount_table.removeRow(selected_row)


class DualExcelViewer(QWidget):
    def __init__(self, leftFilePath, rightFilePath):
        super().__init__()
        self.initUI(leftFilePath, rightFilePath)

    def initUI(self, leftFilePath, rightFilePath):
        self.layout = QVBoxLayout()

        self.splitter = QSplitter(Qt.Horizontal)
        self.splitter.setStyleSheet("QSplitter::handle { background-color: gray; }")

        self.rightViewer = ExcelTreeView(rightFilePath)  # 매출 부분을 ExcelTreeView로 변경
        self.leftViewer = ExcelViewer(leftFilePath, '매입', self.rightViewer)

        self.splitter.addWidget(self.leftViewer)
        self.splitter.addWidget(self.rightViewer)

        self.splitter.splitterMoved.connect(self.adjustFontSize)

        self.layout.addWidget(self.splitter)
        self.setLayout(self.layout)

        self.adjustFontSize()

    def adjustFontSize(self):
        leftWidth = self.splitter.widget(0).width()
        rightWidth = self.splitter.widget(1).width()

        leftFontSize = max(25, leftWidth // 25)
        rightFontSize = max(25, rightWidth // 25)

        self.leftViewer.labelWidget.setFont(self.setFontSize(leftFontSize))
        self.rightViewer.labelWidget.setFont(self.setFontSize(rightFontSize))

    def setFontSize(self, size):
        font = QtGui.QFont()
        font.setPointSize(size)
        return font

class MarketResearchPage2(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

        self.df = None
        self.excel_data = None
        self.not_filtered_df = pd.DataFrame()
        self.load_excel()

    def initUI(self):
        self.layout = QVBoxLayout()

        self.topLayout = QHBoxLayout()

        self.homeButton = ClickableLabel(self)
        home_pixmap = QPixmap("./home.png")  # 이미지 파일 경로 설정
        home_pixmap = home_pixmap.scaled(50, 50, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)  # 적절한 크기로 조정
        self.homeButton.setPixmap(home_pixmap)
        self.homeButton.setAlignment(QtCore.Qt.AlignLeft)

        self.label = QLabel("시장조사 페이지", self)
        font = QtGui.QFont()
        font.setPointSize(25)
        self.label.setFont(font)
        self.label.setAlignment(QtCore.Qt.AlignLeft)

        self.date_label = QLabel(self)
        current_date = QDate.currentDate().toString('yyyy-MM-dd')
        self.date_label.setText(current_date)
        self.date_label.setAlignment(QtCore.Qt.AlignRight)

        self.topLayout.addWidget(self.homeButton)
        self.topLayout.addWidget(self.label)
        self.topLayout.addWidget(self.date_label)

        self.search_layout = QHBoxLayout()
        self.lineEdit = QLineEdit(self)
        self.lineEdit.setPlaceholderText("음식 이름")
        self.lineEdit.setFixedHeight(60)  # 높이 조정
        self.search_button = QPushButton("검색", self)
        self.search_button.setFixedHeight(60)  # 높이 조정
        self.search_button.setFixedWidth(150)  # 폭 조정

        self.lineEdit.setStyleSheet("font-size: 20px;")
        self.lineEdit.returnPressed.connect(self.search_food)
        self.search_button.clicked.connect(self.search_food)

        self.search_layout.addWidget(self.lineEdit)
        self.search_layout.addWidget(self.search_button)

        self.search_layout.setContentsMargins(0, 20, 0, 0)  # 위쪽 여백 추가
        self.lineEdit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.search_button.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)

        self.table_widget = QTableWidget(self)
        self.table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.button_layout = QHBoxLayout()
        self.add_button = QPushButton("추가", self)
        self.delete_button = QPushButton("삭제", self)
        self.button_layout.addStretch()
        self.button_layout.addWidget(self.add_button)
        self.button_layout.addWidget(self.delete_button)

        self.layout.addLayout(self.topLayout)
        self.layout.addLayout(self.search_layout)
        self.layout.addWidget(self.table_widget)
        self.layout.addLayout(self.button_layout)

        self.setLayout(self.layout)

        self.add_button.clicked.connect(self.add_new_row)
        self.delete_button.clicked.connect(self.delete_selected_row)

    def load_excel(self):
        file_name = './매입.xlsx'
        try:
            self.excel_data = pd.read_excel(file_name, sheet_name=None)
            self.df = self.excel_data['물품목록list']
            self.display_data(self.df)
        except Exception as e:
            QMessageBox.critical(self, 'Error', str(e))

    def display_data(self, df):
        self.table_widget.setRowCount(df.shape[0])
        self.table_widget.setColumnCount(df.shape[1])
        self.table_widget.setHorizontalHeaderLabels(df.columns)

        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                item = QTableWidgetItem(str(df.iat[row, col]))
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)  # 읽기 전용으로 설정
                self.table_widget.setItem(row, col, item)

    def search_food(self):
        user_input = self.lineEdit.text()
        if user_input and self.df is not None:
            self.filtered_df = self.df[self.df.apply(lambda row: row.astype(str).str.contains(user_input).any(), axis=1)]
            self.sort_input()
            self.not_filtered_df = self.df[~self.df.apply(lambda row: row.astype(str).str.contains(user_input).any(), axis=1)]
            self.display_data(self.filtered_df)
        if user_input == '':
            self.filtered_df = self.df
            self.not_filtered_df = pd.DataFrame()
            self.display_data(self.filtered_df)

    def sort_input(self):
        current_date = datetime.strptime(QDate.currentDate().toString('yyyy-MM-dd'), '%Y-%m-%d')

        def calculate_priority1(row_date):
            if pd.isnull(row_date):
                return 0
            date_diff = (current_date - row_date).days
            if date_diff <= 0:
                return 0
            elif date_diff <= 90:
                return 1
            elif date_diff <= 180:
                return 2
            elif date_diff <= 270:
                return 3
            elif date_diff <= 365:
                return 4
            else:
                return 5

        # 슬라이스된 데이터 프레임을 명시적으로 복사본을 만듭니다.
        self.filtered_df = self.filtered_df.copy()

        self.filtered_df.loc[:, '우선순위1'] = pd.to_datetime(self.filtered_df['일자']).apply(calculate_priority1)

        temp_df = self.filtered_df.groupby('우선순위1')['단가'].rank(method='min', ascending=True).astype(int)
        self.filtered_df = self.filtered_df.assign(우선순위2=temp_df)

        self.filtered_df = self.filtered_df.sort_values(by=['우선순위1', '우선순위2']).reset_index(drop=True)
        self.filtered_df.drop(columns=['우선순위1', '우선순위2'], inplace=True)

    def delete_selected_row(self):
        current_row = self.table_widget.currentRow()
        if current_row >= 0:
            reply = QMessageBox.question(self, '삭제 확인', '정말로 삭제하시겠습니까?',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                self.table_widget.removeRow(current_row)

    def add_new_row(self):
        row_position = self.table_widget.rowCount()
        self.table_widget.insertRow(row_position)

    def save_current_data_to_excel(self):
        file_name = './매입.xlsx'
        rows = self.table_widget.rowCount()
        cols = self.table_widget.columnCount()
        data = []

        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.table_widget.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)

        df = pd.DataFrame(data, columns=[self.table_widget.horizontalHeaderItem(i).text() for i in range(cols)])
        combined_df = pd.concat([df, self.not_filtered_df])
        self.df = combined_df
        self.excel_data['물품목록list'] = combined_df
        with pd.ExcelWriter(file_name) as writer:
            for sheet_name, sheet_df in self.excel_data.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

class MarketResearchPage(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

        self.df = None
        self.excel_data = None
        self.not_filtered_df = pd.DataFrame()
        self.load_excel()

    def initUI(self):
        self.layout = QVBoxLayout()

        self.topLayout = QHBoxLayout()

        self.homeButton = ClickableLabel(self)
        home_pixmap = QPixmap("./home.png")  # 이미지 파일 경로 설정
        home_pixmap = home_pixmap.scaled(50, 50, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)  # 적절한 크기로 조정
        self.homeButton.setPixmap(home_pixmap)
        self.homeButton.setAlignment(QtCore.Qt.AlignLeft)

        self.label = QLabel("시장조사 페이지", self)
        font = QtGui.QFont()
        font.setPointSize(25)
        self.label.setFont(font)
        self.label.setAlignment(QtCore.Qt.AlignLeft)

        self.date_label = QLabel(self)
        current_date = QDate.currentDate().toString('yyyy-MM-dd')
        self.date_label.setText(current_date)
        self.date_label.setAlignment(QtCore.Qt.AlignRight)

        self.topLayout.addWidget(self.homeButton)
        self.topLayout.addWidget(self.label)
        self.topLayout.addWidget(self.date_label)

        self.search_layout = QHBoxLayout()
        self.lineEdit = QLineEdit(self)
        self.lineEdit.setPlaceholderText("음식 이름")
        self.lineEdit.setFixedHeight(60)  # 높이 조정
        self.search_button = QPushButton("검색", self)
        self.search_button.setFixedHeight(60)  # 높이 조정
        self.search_button.setFixedWidth(150)  # 폭 조정

        self.lineEdit.setStyleSheet("font-size: 20px;")
        self.lineEdit.returnPressed.connect(self.search_food)
        self.search_button.clicked.connect(self.search_food)

        self.search_layout.addWidget(self.lineEdit)
        self.search_layout.addWidget(self.search_button)

        self.search_layout.setContentsMargins(0, 20, 0, 0)  # 위쪽 여백 추가
        self.lineEdit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.search_button.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)

        self.table_widget = QTableWidget(self)
        self.table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.button_layout = QHBoxLayout()
        self.add_button = QPushButton("추가", self)
        self.delete_button = QPushButton("삭제", self)
        self.button_layout.addStretch()
        self.button_layout.addWidget(self.add_button)
        self.button_layout.addWidget(self.delete_button)

        self.layout.addLayout(self.topLayout)
        self.layout.addLayout(self.search_layout)
        self.layout.addWidget(self.table_widget)
        self.layout.addLayout(self.button_layout)

        self.setLayout(self.layout)

        self.add_button.clicked.connect(self.add_new_row)
        self.delete_button.clicked.connect(self.delete_selected_row)

    def load_excel(self):
        file_name = './매입.xlsx'
        try:
            self.excel_data = pd.read_excel(file_name, sheet_name=None)
            self.df = self.excel_data['물품목록list']
            self.display_data(self.df)
        except Exception as e:
            QMessageBox.critical(self, 'Error', str(e))

    def display_data(self, df):
        self.table_widget.setRowCount(df.shape[0])
        self.table_widget.setColumnCount(df.shape[1])
        self.table_widget.setHorizontalHeaderLabels(df.columns)

        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                item = QTableWidgetItem(str(df.iat[row, col]))
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)  # 읽기 전용으로 설정
                self.table_widget.setItem(row, col, item)

    def search_food(self):
        user_input = self.lineEdit.text()
        if user_input and self.df is not None:
            self.filtered_df = self.df[self.df.apply(lambda row: row.astype(str).str.contains(user_input).any(), axis=1)]
            self.sort_input()
            self.not_filtered_df = self.df[~self.df.apply(lambda row: row.astype(str).str.contains(user_input).any(), axis=1)]
            self.display_data(self.filtered_df)
        if user_input == '':
            self.filtered_df = self.df
            self.not_filtered_df = pd.DataFrame()
            self.display_data(self.filtered_df)

    def sort_input(self):
        current_date = datetime.strptime(QDate.currentDate().toString('yyyy-MM-dd'), '%Y-%m-%d')

        def calculate_priority1(row_date):
            if pd.isnull(row_date):
                return 0
            date_diff = (current_date - row_date).days
            if date_diff <= 0:
                return 0
            elif date_diff <= 90:
                return 1
            elif date_diff <= 180:
                return 2
            elif date_diff <= 270:
                return 3
            elif date_diff <= 365:
                return 4
            else:
                return 5

        # 슬라이스된 데이터 프레임을 명시적으로 복사본을 만듭니다.
        self.filtered_df = self.filtered_df.copy()

        self.filtered_df.loc[:, '우선순위1'] = pd.to_datetime(self.filtered_df['일자']).apply(calculate_priority1)

        temp_df = self.filtered_df.groupby('우선순위1')['단가'].rank(method='min', ascending=True).astype(int)
        self.filtered_df = self.filtered_df.assign(우선순위2=temp_df)

        self.filtered_df = self.filtered_df.sort_values(by=['우선순위1', '우선순위2']).reset_index(drop=True)
        self.filtered_df.drop(columns=['우선순위1', '우선순위2'], inplace=True)

    def delete_selected_row(self):
        current_row = self.table_widget.currentRow()
        if current_row >= 0:
            reply = QMessageBox.question(self, '삭제 확인', '정말로 삭제하시겠습니까?',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                self.table_widget.removeRow(current_row)

    def add_new_row(self):
        row_position = self.tableWidget.rowCount()
        self.tableWidget.insertRow(row_position)

        # '종류' 열의 위치를 찾습니다.
        column_index = self.tableWidget.horizontalHeaderItem(self.df.columns.get_loc('종류')).column()

        # 새 행의 '종류' 열에 '시장'을 추가합니다.
        self.tableWidget.setItem(row_position, column_index, QTableWidgetItem("시장"))

        # 나머지 열에 빈 항목을 추가합니다.
        for col in range(self.tableWidget.columnCount()):
            if col != column_index:  # '종류' 열이 아닌 경우
                self.tableWidget.setItem(row_position, col, QTableWidgetItem(""))

    def save_current_data_to_excel(self):
        file_name = './매입.xlsx'
        rows = self.table_widget.rowCount()
        cols = self.table_widget.columnCount()
        data = []

        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.table_widget.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)

        df = pd.DataFrame(data, columns=[self.table_widget.horizontalHeaderItem(i).text() for i in range(cols)])
        combined_df = pd.concat([df, self.not_filtered_df])
        self.df = combined_df
        self.excel_data['물품목록list'] = combined_df
        with pd.ExcelWriter(file_name) as writer:
            for sheet_name, sheet_df in self.excel_data.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
class ClickableLabel(QtWidgets.QLabel):
    clicked = QtCore.pyqtSignal()

    def mouseReleaseEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton:
            self.clicked.emit()


class Ui_StackedWidget(object):
    def setupUi(self, StackedWidget):
        StackedWidget.setObjectName("StackedWidget")
        StackedWidget.resize(1200, 900)

        self.page = QtWidgets.QWidget()
        self.page.setObjectName("page")

        self.imageLabel = QLabel(self.page)
        pixmap = QPixmap("./bapu_label.png")
        pixmap = pixmap.scaled(800, 500, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        self.imageLabel.setPixmap(pixmap)
        self.imageLabel.setAlignment(QtCore.Qt.AlignCenter)

        self.homeButton1 = ClickableLabel(self.page)
        home_pixmap = QPixmap("./home.png")
        home_pixmap = home_pixmap.scaled(50, 50, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        self.homeButton1.setPixmap(home_pixmap)
        self.homeButton1.setAlignment(QtCore.Qt.AlignLeft)
        self.homeButton1.clicked.connect(lambda: StackedWidget.setCurrentIndex(0))

        # self.toolButton = QtWidgets.QToolButton(self.page)
        # font = QtGui.QFont()
        # font.setPointSize(25)
        # self.toolButton.setFont(font)
        # self.toolButton.setStyleSheet("color:white; background-color:rgb(230,100,30);")
        # self.toolButton.setObjectName("toolButton")
        # self.toolButton.setText("매입매출")

        # self.toolButton_2 = QtWidgets.QToolButton(self.page)
        # self.toolButton_2.setFont(font)
        # self.toolButton_2.setObjectName("toolButton_2")
        # self.toolButton_2.setText("원가율")

        # self.toolButton_4 = QtWidgets.QToolButton(self.page)
        # self.toolButton_4.setFont(font)
        # self.toolButton_4.setObjectName("toolButton_3")
        # self.toolButton_4.setText("시장조사")

        # self.toolButton_3 = QtWidgets.QToolButton(self.page)
        # self.toolButton_3.setFont(font)
        # self.toolButton_3.setObjectName("toolButton_4")
        # self.toolButton_3.setText("발주서")

        self.toolButton = self.createImageButton(self.page, "./apple.png", "매입매출",
                                                 lambda: StackedWidget.setCurrentIndex(1), 'white',
                                                 text_margin=(0, 30, 0, 0))
        self.toolButton_2 = self.createImageButton(self.page, "./eggplant.png", "원가율",
                                                   lambda: StackedWidget.setCurrentIndex(2), 'white',
                                                   text_margin=(0, 50, -20, 0))
        self.toolButton_3 = self.createImageButton(self.page, "./banana.png", "발주서",
                                                   lambda: StackedWidget.setCurrentIndex(3), 'black',
                                                   text_margin=(0, 60, 0, 0))
        self.toolButton_4 = self.createImageButton(self.page, "./carrot.png", "시장조사",
                                                   lambda: StackedWidget.setCurrentIndex(4), 'white',
                                                   text_margin=(0, 40, -40, 0))

        # self, parent, imagePath, text, clickHandler, text_color, text_margin=(0, 30, 0, 0)
        self.pageLayout = QtWidgets.QVBoxLayout(self.page)
        self.pageLayout.addWidget(self.homeButton1, alignment=QtCore.Qt.AlignLeft)
        self.pageLayout.addWidget(self.imageLabel, alignment=QtCore.Qt.AlignCenter)

        spacer = QtWidgets.QSpacerItem(0, 0, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.pageLayout.addItem(spacer)

        self.buttonLayout = QHBoxLayout()
        self.buttonLayout.addWidget(self.toolButton)
        self.buttonLayout.addWidget(self.toolButton_2)
        self.buttonLayout.addWidget(self.toolButton_3)
        self.buttonLayout.addWidget(self.toolButton_4)

        self.pageLayout.addLayout(self.buttonLayout)
        StackedWidget.addWidget(self.page)

        self.setupPurchaseSalePage(StackedWidget)
        self.setupCostRatePage(StackedWidget)
        self.setupOrderFormPage(StackedWidget)
        self.setupMarketResearchPage(StackedWidget)

        self.toolButton.clicked.connect(lambda: StackedWidget.setCurrentIndex(1))
        self.toolButton_2.clicked.connect(lambda: StackedWidget.setCurrentIndex(2))
        self.toolButton_3.clicked.connect(lambda: StackedWidget.setCurrentIndex(3))
        self.toolButton_4.clicked.connect(lambda: StackedWidget.setCurrentIndex(4))

    def createImageButton(self, parent, imagePath, text, clickHandler, text_color, text_margin=(0, 30, 0, 0)):
        label = ClickableLabel(parent)
        pixmap = QPixmap(imagePath)
        pixmap = pixmap.scaled(180, 180, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        label.setPixmap(pixmap)
        label.setAlignment(QtCore.Qt.AlignCenter)

        layout = QVBoxLayout(label)
        layout.setContentsMargins(*text_margin)  # 위쪽 여백을 10으로 설정하여 라벨을 아래로 이동
        textLabel = QLabel(text)
        textLabel.setAlignment(QtCore.Qt.AlignCenter)
        textLabel.setFont(QFont('Arial', 30))
        textLabel.setStyleSheet(f"color: {text_color};")
        layout.addWidget(textLabel)

        label.clicked.connect(clickHandler)
        return label

    def setupMarketResearchPage(self, StackedWidget):
        self.page_3 = QtWidgets.QWidget()
        self.page_3.setObjectName("page_3")

        self.market_research_page = MarketResearchPage()
        self.market_research_page.homeButton.clicked.connect(lambda: StackedWidget.setCurrentIndex(0))

        self.page3Layout = QVBoxLayout(self.page_3)
        self.page3Layout.addWidget(self.market_research_page)

        StackedWidget.addWidget(self.page_3)

    def setupPurchaseSalePage(self, StackedWidget):
        self.page_2 = QtWidgets.QWidget()
        self.page_2.setObjectName("page_2")

        self.homeButton2 = QtWidgets.QLabel(self.page_2)
        home_pixmap = QPixmap("./home.png")
        home_pixmap = home_pixmap.scaled(50, 50, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        self.homeButton2 = ClickableLabel(self.page_2)
        self.homeButton2.setPixmap(home_pixmap)
        self.homeButton2.setAlignment(QtCore.Qt.AlignLeft)
        self.homeButton2.setFixedHeight(50)
        self.homeButton2.clicked.connect(lambda: StackedWidget.setCurrentIndex(0))
        self.page2Layout = QtWidgets.QVBoxLayout(self.page_2)
        self.page2Layout.addWidget(self.homeButton2, alignment=QtCore.Qt.AlignLeft)

        self.dualExcelViewer = DualExcelViewer(data_path_purchase, data_path_sales)
        self.page2Layout.addWidget(self.dualExcelViewer)
        StackedWidget.addWidget(self.page_2)

    def setupCostRatePage(self, StackedWidget):
        self.page_4 = QtWidgets.QWidget()
        self.page_4.setObjectName("page_4")

        self.homeButton4 = QtWidgets.QLabel(self.page_4)
        home_pixmap = QPixmap("./home.png")
        home_pixmap = home_pixmap.scaled(50, 50, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        self.homeButton4 = ClickableLabel(self.page_4)
        self.homeButton4.setPixmap(home_pixmap)
        self.homeButton4.setAlignment(QtCore.Qt.AlignLeft)
        self.homeButton4.clicked.connect(lambda: StackedWidget.setCurrentIndex(0))

        self.label_5 = QtWidgets.QLabel(self.page_4)
        font = QtGui.QFont()
        font.setPointSize(40)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_5.setText("원가율 페이지")

        self.page4Layout = QtWidgets.QVBoxLayout(self.page_4)
        self.page4Layout.addWidget(self.homeButton4, alignment=QtCore.Qt.AlignLeft)
        self.page4Layout.addWidget(self.label_5, alignment=QtCore.Qt.AlignCenter)

        self.readOnlyExcelViewer = ReadOnlyExcelViewer(data_path_cost_ratio, '원가율')
        self.page4Layout.addWidget(self.readOnlyExcelViewer)
        StackedWidget.addWidget(self.page_4)

    def setupOrderFormPage(self, StackedWidget):
        self.page_5 = QtWidgets.QWidget()
        self.page_5.setObjectName("page_5")

        self.orderFormViewer = OrderFormViewer()
        self.orderFormViewer.homeButton.clicked.connect(lambda: StackedWidget.setCurrentIndex(0))

        self.page5Layout = QVBoxLayout(self.page_5)
        self.page5Layout.addWidget(self.orderFormViewer)

        StackedWidget.addWidget(self.page_5)

    def retranslateUi(self, StackedWidget):
        _translate = QtCore.QCoreApplication.translate
        StackedWidget.setWindowTitle(_translate("StackedWidget", "BapuriFNB"))


class MyStackedWidget(QtWidgets.QStackedWidget):
    def closeEvent(self, event):
        reply = QtWidgets.QMessageBox.question(self, '종료 확인', '종료 하시겠습니까?',
                                               QtWidgets.QMessageBox.Yes |
                                               QtWidgets.QMessageBox.No,
                                               QtWidgets.QMessageBox.Yes)
        if reply == QtWidgets.QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    StackedWidget = MyStackedWidget()
    ui = Ui_StackedWidget()
    ui.setupUi(StackedWidget)
    StackedWidget.show()
    sys.exit(app.exec_())
